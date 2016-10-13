import os

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from country.models import Country
from country.utils import update_countries
from taxon.models import Taxon
from taxon.utils import update_taxonomy
from .forms import ProjectLineForm
from .services import limsfm_get_project


SAMPLE_SHEET_COL_ORDER = [
    'well_alpha',
    'sample_ref',
    'customers_ref',
    'dna_concentration_ng_ul',
    'volume_ul',
    'taxon_name',
    'collection_day',
    'collection_month',
    'collection_year',
    'geo_country_name',
    'geo_specific_location',
    'lab_experiment_type',
    'lab_further_details',
    'host_taxon_name',
    'host_sample_type',
    'host_further_details',
    'environmental_sample_type',
    'env_further_details',
]


def create_sample_sheet(project_uuid):
    """Create sample sheet with 'initial data' for project"""
    wb = load_workbook(os.path.join(
        os.path.dirname(__file__),
        'static/portal/excel/mng_excel_template.xlsx'))
    ws = wb['Data']

    project = limsfm_get_project(project_uuid)

    ws['F1'] = project['reference']
    data_rows = ws.rows[6:]

    for i, pl in enumerate(project['projectlines']):
        row = data_rows[i]

        row[1].value = pl['sample_ref']

        if pl['customers_ref']:
            row[2].value = pl['customers_ref']
            if pl['dna_concentration_ng_ul']:
                row[3].value = float(pl['dna_concentration_ng_ul'])
            if pl['volume_ul']:
                row[4].value = float(pl['volume_ul'])
            row[5].value = pl['taxon_name']
            if pl['collection_day']:
                row[6].value = int(pl['collection_day'])
            if pl['collection_month']:
                row[7].value = int(pl['collection_month'])
            if pl['collection_year']:
                row[8].value = int(pl['collection_year'])
            row[9].value = pl['geo_country_name']
            row[10].value = pl['geo_specific_location']

            if pl['study_type'] == "Lab":
                row[11].value = pl['lab_experiment_type']
                row[12].value = pl['further_details']
            elif pl['study_type'] == "Host":
                row[13].value = pl['host_taxon_name']
                row[14].value = pl['host_sample_type']
                row[15].value = pl['further_details']
            elif pl['study_type'] == "Environmental":
                row[16].value = pl['environmental_sample_type']
                row[17].value = pl['further_details']

    # Taxon list validation
    taxon_validator = DataValidation(type='list', formula1='=Lookups!$A:$A',
                                     allow_blank=True)
    taxon_validator.error = "Please select a taxon from the list"
    taxon_validator.errorTitle = "Invalid taxon"
    ws.add_data_validation(taxon_validator)
    taxon_validator.ranges.append('F7:F102')
    taxon_validator.ranges.append('N7:N102')

    # Country list validation
    country_validator = DataValidation(
        type='list', formula1='=Lookups!$B$1:$B$267', allow_blank=True)
    country_validator.error = "Please select a country from the list"
    country_validator.errorTitle = "Invalid country"
    ws.add_data_validation(country_validator)
    country_validator.ranges.append('J7:J102')

    # Environmental sample type list validation
    est_validator = DataValidation(
        type='list', formula1='=Lookups!$C$1:$C$31', allow_blank=True)
    est_validator.error = ("Please select a environmental sample type from "
                           "the list")
    est_validator.errorTitle = "Invalid environmental sample type"
    ws.add_data_validation(est_validator)
    est_validator.ranges.append('Q7:Q102')

    # Host sample type list validation
    hst_validator = DataValidation(
        type='list', formula1='=Lookups!$D$1:$D$67', allow_blank=True)
    hst_validator.error = "Please select a host sample type from the list"
    hst_validator.errorTitle = "Invalid host sample type"
    ws.add_data_validation(hst_validator)
    hst_validator.ranges.append('O7:O102')

    # Lab experiment type
    lab_validator = DataValidation(
        type='list', formula1='=Lookups!$E$1:$E$5', allow_blank=True)
    lab_validator.error = "Please select a host sample type from the list"
    lab_validator.errorTitle = "Invalid host sample type"
    ws.add_data_validation(lab_validator)
    lab_validator.ranges.append('L7:L102')

    return wb


def parse_sample_sheet(project, sheet):
    """Parse uploaded sample sheet data; return errors and updates"""

    projectlines = {}
    for pl in project['projectlines']:
        projectlines[pl['sample_ref']] = pl

    updates = {}
    errors = []

    for row in sheet.row_range()[6:102]:
        row_data = dict(zip(SAMPLE_SHEET_COL_ORDER, sheet.row[row]))
        sample_ref = str(row_data['sample_ref'])

        # Skip 'blank' rows
        if not row_data['customers_ref']:
            continue

        # Check sample ref in project
        if sample_ref not in projectlines:
            errors.append(
                {'row': row, 'message': "Barcode ID %(barcode)s is not "
                 "part of this project." %
                 {'row': row, 'barcode': sample_ref}})
            continue

        # Check only one meta data section filled in
        has_lab_meta = (bool(row_data['lab_experiment_type']) or
                        bool(row_data['lab_further_details']))
        has_host_meta = (bool(row_data['host_taxon_name']) or
                         bool(row_data['host_sample_type']) or
                         bool(row_data['host_further_details']))
        has_env_meta = (bool(row_data['environmental_sample_type']) or
                        bool(row_data['env_further_details']))
        if has_lab_meta + has_host_meta + has_env_meta != 1:
            errors.append(
                {'row': row, 'message': " Meta data is required in ONE "
                 "of the three coloured sections."})
            continue

        # Set study_type, further details
        if has_lab_meta:
            row_data['study_type'] = 'Lab'
            row_data['further_details'] = row_data['lab_further_details']
        elif has_host_meta:
            row_data['study_type'] = 'Host'
            row_data['further_details'] = row_data['host_further_details']
        elif has_env_meta:
            row_data['study_type'] = 'Environmental'
            row_data['further_details'] = row_data['env_further_details']

        # Pass row data to form
        row_data['aliquottype_name'] = (projectlines[sample_ref]
                                        ['aliquottype_name'])
        form = ProjectLineForm(row_data)
        if form.is_valid():
            uuid = projectlines[sample_ref]['uuid']
            updates[uuid] = form.cleaned_data
        else:
            for i, col in enumerate(form.errors):
                errors.append(
                    {'row': row, 'message': form.errors[col][0]})

    if not (errors or updates):
        errors.append({'row': 0, 'message': "No rows to update."})

    return {'errors': errors, 'updates': updates}


def update_sample_sheet_template():
    """Update the sample sheet template (add lookup values)"""
    update_countries()
    update_taxonomy()

    wb = load_workbook(os.path.join(
        os.path.dirname(__file__),
        'static/portal/excel/mng_excel_template_base.xlsx'))
    ws = wb['Lookups']

    print("Updating excel template taxonomy lookup")
    taxonomy = Taxon.objects.all().order_by('name')
    for r, taxon in enumerate(taxonomy):
        ws.cell(row=r + 1, column=1).value = taxon.name

    print("Updating excel template country lookup")
    countries = Country.objects.all().order_by('name')
    for r, country in enumerate(countries):
        ws.cell(row=r + 1, column=2).value = country.name

    wb.save(os.path.join(
        os.path.dirname(__file__),
        'static/portal/excel/mng_excel_template.xlsx'))
